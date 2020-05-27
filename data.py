import openpyxl as xl
import numpy as np

workbook = xl.load_workbook('bc_data.xlsx')
sheet = workbook['breast-cancer-wisconsin (1)']

# total 699 data
# 500 for training
first_row_of_train_data = 2
last_row_of_train_data = 500
train_data = []
train_labels = []
for row in range(first_row_of_train_data, last_row_of_train_data + 1):
    train_data.append([])
    '''for train labels'''
    train_labels.append(sheet.cell(row, 11).value)

#print(f'the length of empty train_data : {len(train_data)}')
#print(f'The train_labels length is :{len(train_labels)}\n train_labels :{train_labels}')

for column in range(2, 11):
    for row_number, z in zip(range(first_row_of_train_data, last_row_of_train_data + 1),
                             range(last_row_of_train_data - 1)):  # change row num here
        train_data[z].append(sheet.cell(int(row_number), column).value)
#print(f'train_data : {train_data}')

train_data_arr = np.array(train_data)
print(f'train_data array : {train_data_arr}')
train_data_arr_final = train_data_arr.reshape(train_data_arr.shape[0], 1, 9)

train_labels_arr = np.array(train_labels).T
#print(f'train_labels array: {train_labels_arr}')
train_labels_arr_final = train_labels_arr.reshape(train_data_arr.shape[0], 1, 1)

'''for test data'''
first_row_of_test_data = 500
last_row_of_test_data = 700
test_data = []
test_labels = []
for row in range(first_row_of_test_data, last_row_of_test_data + 1):
    test_data.append([])
    '''for train labels'''
    test_labels.append(sheet.cell(row, 11).value)
#print(f'The test_data length: {len(test_data)}\nThe test_labels: {test_labels}')
for column in range(2, 11):
    for row_number, z in zip(range(first_row_of_test_data, last_row_of_test_data + 1),
                             range(last_row_of_test_data - 1)):  # change row num here
        test_data[z].append(sheet.cell(int(row_number), column).value)
#print(f'test_data : {test_data}')

test_data_arr = np.array(test_data)
test_data_arr_final = test_data_arr.reshape(test_data_arr.shape[0], 1, 9)
print(f'the test_data array is: {test_data_arr_final}')

test_labels_arr = np.array(test_labels)
#print(f'test_labels array: {test_labels_arr}')
test_labels_arr_final = test_labels_arr.reshape(test_data_arr.shape[0], 1, 1)